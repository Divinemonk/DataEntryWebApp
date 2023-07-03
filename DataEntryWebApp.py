from flask import Flask, render_template, request, redirect, url_for
from markupsafe import Markup
from openpyxl import load_workbook, Workbook
import webbrowser
import sys

app = Flask(__name__, static_folder='static', static_url_path='/static')

# Store the filename in a global variable
filename = 'data.xlsx'

# Home route to display the form
@app.route('/')
def home():
    default_filename = filename  # Use the global filename variable as the default filename
    return render_template('form.html', filename=default_filename, default_filename=default_filename)

# Setting up favicon
@app.route('/favicon.ico')
def favicon():
    return app.send_static_file('assets/data_entry_logo.ico')

@app.context_processor
def inject_url_for():
    return dict(url_for=url_for)

# Route to process the filename form submission
@app.route('/submit_filename', methods=['POST'])
def submit_filename():
    global filename  # Access the global filename variable
    # Retrieve the filename from the form
    filename = request.form['filename']

    # Display success notification
    notification = {
        'type': 'success',
        'message': 'Filename has been saved successfully'
    }

    return render_template('form.html', filename=filename, notification=notification)

# Route to process the data entry form submission
@app.route('/submit_data', methods=['POST'])
def submit_data():
    # Retrieve the form data
    name = request.form['name']
    gender = request.form['gender']
    email = request.form['email']
    address = request.form['address']
    role = request.form['role']
    
    try:
        # Load the existing workbook or create a new one using the global filename variable
        try:
            workbook = load_workbook(filename)
        except FileNotFoundError:
            workbook = Workbook()
            
        # Get the active sheet
        sheet = workbook.active

        # Add headers if the sheet is empty
        if sheet.max_row == 1:
            sheet['A1'] = 'ID'
            sheet['B1'] = 'Name'
            sheet['C1'] = 'Gender'
            sheet['D1'] = 'Email'
            sheet['E1'] = 'Address'
            sheet['F1'] = 'Role'

        # Find the next ID value
        next_id = sheet.cell(row=sheet.max_row, column=1).value + 1 if sheet.max_row > 1 else 1

        # Populate the row with the form data
        sheet.cell(row=sheet.max_row + 1, column=1, value=next_id)
        sheet.cell(row=sheet.max_row, column=2, value=name)
        sheet.cell(row=sheet.max_row, column=3, value=gender)
        sheet.cell(row=sheet.max_row, column=4, value=email)
        sheet.cell(row=sheet.max_row, column=5, value=address)
        sheet.cell(row=sheet.max_row, column=6, value=role)

        # Save the workbook
        workbook.save(filename)

        # Display success notification
        notification = {
            'type': 'success',
            'message': 'Data has been saved successfully'
        }
    except Exception as e:
        # Display error notification
        notification = {
            'type': 'error',
            'message': Markup(f'An error occurred: {str(e)}')
        }

    return render_template('form.html', filename=filename, notification=notification)

# Route to display the data in tabular form and allow deletion and editing
@app.route('/view_data')
def view_data():
    try:
        # Load the workbook
        workbook = load_workbook(filename)
        
        # Get the active sheet
        sheet = workbook.active
        
        # Get all the rows as a list
        rows = list(sheet.iter_rows(values_only=True))
        
        # Exclude the header row
        data_rows = rows[1:]
        
        return render_template('view_data.html', data=data_rows)
    except FileNotFoundError:
        # Display error notification if the file is not found
        notification = {
            'type': 'error',
            'message': Markup(f'File "{filename}" not found')
        }
        return render_template('form.html', filename=filename, notification=notification)
    except Exception as e:
        # Display error notification for other exceptions
        notification = {
            'type': 'error',
            'message': Markup(f'An error occurred: {str(e)}')
        }
        return render_template('form.html', filename=filename, notification=notification)

# Route to delete a row of data
@app.route('/delete_data/<int:row_id>', methods=['POST'])
def delete_data(row_id):
    try:
        # Load the workbook
        workbook = load_workbook(filename)

        # Get the active sheet
        sheet = workbook.active

        # Delete the row by shifting all rows below it up
        sheet.delete_rows(row_id + 1)

        # Update the ID values
        for row in sheet.iter_rows(min_row=2, min_col=1, values_only=True):
            if row[0] > row_id:
                sheet.cell(row=row[0], column=1, value=row[0] - 1)

        # Save the workbook
        workbook.save(filename)

        # Redirect back to the view_data route
        return redirect('/view_data')
    except FileNotFoundError:
        # Display error notification if the file is not found
        notification = {
            'type': 'error',
            'message': Markup(f'File "{filename}" not found')
        }
        return render_template('form.html', filename=filename, notification=notification)
    except Exception as e:
        # Display error notification for other exceptions
        notification = {
            'type': 'error',
            'message': Markup(f'An error occurred: {str(e)}')
        }
        return render_template('form.html', filename=filename, notification=notification)

# Route to edit a row of data
@app.route('/edit_data/<int:row_id>', methods=['GET', 'POST'])
def edit_data(row_id):
    try:
        # Load the workbook
        workbook = load_workbook(filename)
        
        # Get the active sheet
        sheet = workbook.active
        
        if request.method == 'GET':
            # Get the row to edit
            row = list(sheet.iter_rows(min_row=row_id + 1, max_row=row_id + 1, values_only=True))[0]
            
            return render_template('edit_data.html', row=row, row_id=row_id)
        elif request.method == 'POST':
            # Retrieve the form data
            name = request.form['name']
            gender = request.form['gender']
            email = request.form['email']
            address = request.form['address']
            role = request.form['role']
            
            # Update the row with the form data
            sheet.cell(row=row_id + 1, column=2, value=name)
            sheet.cell(row=row_id + 1, column=3, value=gender)
            sheet.cell(row=row_id + 1, column=4, value=email)
            sheet.cell(row=row_id + 1, column=5, value=address)
            sheet.cell(row=row_id + 1, column=6, value=role)
            
            # Save the workbook
            workbook.save(filename)
            
            # Redirect back to the view_data route
            return redirect('/view_data')
    except FileNotFoundError:
        # Display error notification if the file is not found
        notification = {
            'type': 'error',
            'message': Markup(f'File "{filename}" not found')
        }
        return render_template('form.html', filename=filename, notification=notification)
    except Exception as e:
        # Display error notification for other exceptions
        notification = {
            'type': 'error',
            'message': Markup(f'An error occurred: {str(e)}')
        }
        return render_template('form.html', filename=filename, notification=notification)

# Route to display the modify fields form
@app.route('/modify_fields')
def modify_fields():
    fields = get_fields()
    return render_template('modify_fields.html', fields=fields)

# Route to process the modify fields form submission
@app.route('/modify_fields', methods=['POST'])
def submit_modified_fields():
    new_fields = [request.form['field1'], request.form['field2'], request.form['field3']]
    save_fields(new_fields)

    # Display success notification
    notification = {
        'type': 'success',
        'message': 'Input fields have been modified successfully'
    }

    return render_template('modify_fields.html', fields=new_fields, notification=notification)

# Helper functions to save and retrieve the modified input fields
def save_fields(fields):
    # Save the fields to a file, database, or any other storage method
    pass

def get_fields():
    # Retrieve the fields from the saved file, database, or storage method
    # Return the fields as a list
    return ['Field 1', 'Field 2', 'Field 3']  # Replace with your logic to retrieve the fields


if __name__ == '__main__':
    # Open the default browser with the specified URL
    webbrowser.open('http://localhost:7777')

    if sys.platform.startswith('win'):
        # Run the Flask app using Waitress as the production server on Windows
        from waitress import serve
        serve(app, host='0.0.0.0', port=7777)
    else:
        # Run the Flask app using Gunicorn as the production server on Unix
        from gunicorn.app.base import BaseApplication

        class FlaskApplication(BaseApplication):
            def __init__(self, app, options=None):
                self.options = options or {}
                self.application = app
                super(FlaskApplication, self).__init__()

            def load_config(self):
                for key, value in self.options.items():
                    if key in self.cfg.settings and value is not None:
                        self.cfg.set(key.lower(), value)

            def load(self):
                return self.application

        options = {
            'bind': '0.0.0.0:7777',  # Bind the server to all network interfaces on port 7777
            'workers': 4,  # Number of worker processes
            'threads': 2,  # Number of threads per worker process
            'worker_class': 'sync',  # Worker class (synchronous worker)
            'worker_connections': 1000,  # Maximum number of simultaneous connections
            'timeout': 30,  # Timeout for requests
            'keepalive': 2  # Number of seconds to keep idle connections alive
        }

        FlaskApplication(app, options).run()
